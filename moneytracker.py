import argparse
import math
import os
import sys
from datetime import datetime, timedelta  # to manipulate dates

import pandas as pd  # to perform data manipulation and analysis
from openpyxl import load_workbook


class Money(object):
    TransactionFile = r"transaction.xlsx"
    TransactionFileSheetName = "Transactions"
    AccountBalanceSheetName = "Account Balance"
    MonthlyCategorySheetName = "Monthly Category"

    def __init__(self):
        self.transactionfile = pd.read_excel(self.TransactionFile, sheet_name=None)
        self.now = datetime.now()
        self.month_year_prev = (self.now.replace(day=1) - timedelta(days=1)).strftime("%b %Y")
        self.month_year_now = self.now.strftime("%b %Y")

    def calculate_share(self):
        transactions = self.transactionfile[self.TransactionFileSheetName]
        last_payment_date = transactions.loc[transactions["Category"] == "Share_Payment"]["Date"].max()
        if pd.isna(last_payment_date):
            last_payment_date = datetime.strptime("2021-05-01", "%Y-%m-%d")
        trans_after = transactions.loc[transactions["Date"] > last_payment_date]
        jayne = trans_after.loc[trans_after["Description"] == "Jayne's"]
        jayne_category = (
                (
                        jayne.groupby("Category")["Amount"].sum()//10
                ).apply(abs).astype(int) * 10
        ).apply(math.ceil)

        shared = trans_after.loc[trans_after["Share"]]
        shared_category = (
                (
                        shared.groupby("Category")["Amount"].sum()//20)
                .apply(abs).astype(int) * 10
        ) \
            .apply(math.ceil)
        category = shared_category.add(jayne_category, fill_value=0)
        print(f"For share payment \n{category.to_markdown()}\n\n")
        print(f"Total payment: {category.sum()}")

    def monthly_category(self):
        transactions = self.transactionfile[self.TransactionFileSheetName]
        transactions["Month_Year"] = transactions["Date"].dt.strftime("%b %Y")
        monthly_summary = self.transactionfile.get(self.MonthlyCategorySheetName, pd.DataFrame([]))
        trans_month_year = transactions.groupby("Month_Year")
        if monthly_summary.empty:
            my = None
        elif self.month_year_now in monthly_summary.columns:
            my = self.month_year_now
        else:
            if self.month_year_now not in monthly_summary.columns and \
                    self.month_year_now in transactions["Month_Year"].values:
                my = self.month_year_now
            else:
                my = self.month_year_prev

        if my is None:
            mt = transactions.groupby("Month_Year")
            for each_my in mt.groups.keys():
                pass

        monthly_summary_columns = []
        if not monthly_summary.empty:
            monthly_summary_columns = monthly_summary.columns[1:]
            for month in monthly_summary_columns:
                try:
                    trans_month_year.pop(month)
                except KeyError:
                    continue

        current_summary = transactions.groupby(by=["Category", "Month_Year"])["Amount"] \
            .sum().reset_index(name="Sum")

    def update_balance(self):
        account_balance = self.transactionfile[self.AccountBalanceSheetName]
        if Money._check_columns(account_balance.columns):
            print(f"Check columns in Account Balance sheet\n{account_balance.columns}")
            sys.exit()
        transactions = self.transactionfile[self.TransactionFileSheetName]

        latest = account_balance.columns[-1]
        current_balance = pd.Series(account_balance[latest].values,
                                    index=account_balance[account_balance.columns[0]].values)
        balance = account_balance.copy()

        if latest == "Init Balance":
            latest = "2021-05-01"
        transactions = transactions.loc[transactions["Date"] > pd.to_datetime(latest)]

        balance[self.now.strftime("%Y-%m-%d")] = [transactions[transactions["Source"] == account]["Amount"].sum() +
                                                  current_balance[account]
                                                  for account in balance["Account"]]

        Money.append_df_to_excel(self.TransactionFile, balance, self.AccountBalanceSheetName, startrow=0, index=False)

    @staticmethod
    def _check_columns(columns):
        for col in columns:
            if col.startswith("Unnamed"):
                return True
        return False

    @staticmethod
    def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        """
        Append a DataFrame [df] to existing Excel file [filename]
        into [sheet_name] Sheet.
        If [filename] doesn't exist, then this function will create it.

        @param filename: File path or existing ExcelWriter
                         (Example: '/path/to/file.xlsx')
        @param df: DataFrame to save to workbook
        @param sheet_name: Name of sheet which will contain DataFrame.
                           (default: 'Sheet1')
        @param startrow: upper left cell row to dump data frame.
                         Per default (startrow=None) calculate the last row
                         in the existing DF and write to the next row...
        @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                               before writing DataFrame to Excel file
        @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                                [can be a dictionary]
        @return: None

        Usage examples:

        >>> append_df_to_excel('d:/temp/test.xlsx', df)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                               index=False)

        >>> append_df_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                               index=False, startrow=25)

        (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
        """
        # Excel file doesn't exist - saving and exiting
        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name,
                startrow=startrow if startrow is not None else 0,
                **to_excel_kwargs)
            return

        # ignore [engine] parameter if it was passed
        if 'engine' in to_excel_kwargs:
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        # write out the new sheet
        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

        # save the workbook
        writer.save()


if __name__ == "__main__":
    arg_parser = argparse.ArgumentParser(description="Money tracker")
    arg_parser.add_argument('-c', '--share', help="Calculate share payment", action="store_true")
    arg_parser.add_argument('-u', '--update-balance', help="Update account balance", action="store_true")
    args = arg_parser.parse_args()
    m = Money()

    if args.share:
        m.calculate_share()
    elif args.update_balance:
        m.update_balance()
